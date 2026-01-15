import csv
import logging
import os
import threading
import weakref
from functools import lru_cache

from PyQt6.QtCore import Qt, QPropertyAnimation, QEasingCurve, QSortFilterProxyModel
from PyQt6.QtGui import QStandardItemModel, QStandardItem, QAction
from PyQt6.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QTableView, QLineEdit, QFormLayout,
    QComboBox, QHeaderView, QSplitter, QTreeView, QMessageBox,
    QGroupBox, QToolBar, QAbstractItemView, QDialog, QInputDialog
)

from models import SQLiteTableModel

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.cidfonts import UnicodeCIDFont
    import qrcode
    import io
    import http.server
    import socketserver
    import threading
    import json
    import urllib.parse

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from data_manager import DataManager, get_category_description, get_app_dir
from ui_theme import AnimatedButton, apply_global_style
from dialogs import PrintLabelsDialog

logger = logging.getLogger(__name__)

# Глобальная переменная для веб-сервера
_web_server_thread = None
_web_server_running = False


class EnhancedProxyModel(QSortFilterProxyModel):
    """Улучшенная прокси-модель с множественной фильтрацией."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.filters = {}
        self.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)

    def set_filter(self, column, text):
        """Установка фильтра для конкретной колонки."""
        if text:
            self.filters[column] = text.lower()
        elif column in self.filters:
            del self.filters[column]
        self.invalidateFilter()

    def clear_filters(self):
        """Очистка всех фильтров."""
        self.filters.clear()
        self.invalidateFilter()

    def filterAcceptsRow(self, source_row, source_parent):
        """Проверка строки на соответствие всем фильтрам."""
        if not self.filters:
            return True
        model = self.sourceModel()
        for column, filter_text in self.filters.items():
            index = model.index(source_row, column, source_parent)
            data = str(model.data(index, Qt.ItemDataRole.DisplayRole) or "").lower()
            if filter_text not in data:
                return False
        return True


class ViewWindow(QMainWindow):
    """Окно просмотра архива с улучшенными фильтрами и поиском."""

    def __init__(self, main_menu=None):
        super().__init__(parent=None)
        self.main_menu = weakref.ref(main_menu) if main_menu else None
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.Window)
        self.setWindowTitle("Просмотр архива")
        self.resize(1500, 850)
        self.manager = DataManager()
        self.db_file = self.manager.db_file
        # Сохранение BASE_URL для использования в QR-кодах
        self.base_url = None
        self._setup_styles()
        self._init_model()
        self._create_actions()
        self._create_menu_bar()
        self._create_tool_bar()
        self._create_ui()
        try:
            self.refresh_data()
        except Exception as e:
            logger.error(f"Ошибка при начальной загрузке: {e}")
            QMessageBox.critical(self, "Ошибка", str(e))
        self._animate_window()

    def _setup_styles(self):
        """Настройка современного стиля."""
        apply_global_style(self)

    def _init_model(self):
        """Инициализация моделей для таблицы и дерева."""
        self.model = SQLiteTableModel(self.manager.conn)
        self.proxy_model = EnhancedProxyModel(self)
        self.proxy_model.setSourceModel(self.model)
        self.tree_model = QStandardItemModel()
        self.tree_model.setHorizontalHeaderLabels(["Иерархия архива"])

    def _create_actions(self):
        """Создание действий для меню."""
        self.refresh_action = QAction("🔄 Обновить", self)
        self.refresh_action.triggered.connect(self.refresh_data)
        self.export_action = QAction("📤 Экспорт в Excel", self)
        self.export_action.triggered.connect(self.export_to_excel)
        self.export_csv_action = QAction("📊 Экспорт в CSV", self)
        self.export_csv_action.triggered.connect(self.export_to_csv)
        self.print_labels_action = QAction("🏷️ Печать наклеек", self)
        self.print_labels_action.triggered.connect(self.print_labels)
        self.migrate_action = QAction("🔄 Миграция из JSON", self)
        self.migrate_action.triggered.connect(self.migrate_data)
        self.back_action = QAction("⬅ Назад в меню", self)
        self.back_action.triggered.connect(self.back_to_menu)

    def _create_menu_bar(self):
        """Создание меню."""
        menu_bar = self.menuBar()
        file_menu = menu_bar.addMenu("📁 Файл")
        file_menu.addAction(self.export_action)
        file_menu.addAction(self.export_csv_action)
        file_menu.addAction(self.print_labels_action)
        file_menu.addAction(self.migrate_action)
        file_menu.addSeparator()
        file_menu.addAction(self.back_action)
        view_menu = menu_bar.addMenu("🔍 Вид")
        view_menu.addAction(self.refresh_action)

    def _create_tool_bar(self):
        """Создание панели инструментов."""
        toolbar = QToolBar("Основная панель")
        self.addToolBar(toolbar)
        refresh_btn = AnimatedButton("🔄 Обновить")
        refresh_btn.clicked.connect(self.refresh_data)
        export_btn = AnimatedButton("📤 Экспорт")
        export_btn.clicked.connect(self.export_to_excel)
        export_csv_btn = AnimatedButton("📊 CSV")
        export_csv_btn.clicked.connect(self.export_to_csv)
        print_labels_btn = AnimatedButton("🏷️ Наклейки")
        print_labels_btn.clicked.connect(self.print_labels)
        back_btn = AnimatedButton("⬅ Назад")
        back_btn.clicked.connect(self.back_to_menu)
        toolbar.addWidget(refresh_btn)
        toolbar.addWidget(export_btn)
        toolbar.addWidget(export_csv_btn)
        toolbar.addWidget(print_labels_btn)
        toolbar.addWidget(back_btn)

    def _create_ui(self):
        """Создание основного интерфейса."""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        filter_group = QGroupBox("🔍 Фильтры")
        filter_layout = QFormLayout()
        self.name_filter = QLineEdit()
        self.name_filter.setPlaceholderText("Поиск по названию...")
        self.name_filter.textChanged.connect(lambda: self.proxy_model.set_filter(1, self.name_filter.text()))
        filter_layout.addRow("Название:", self.name_filter)
        self.type_filter = QComboBox()
        self.type_filter.addItems(["Все", "Документ", "Коробка", "Папка", "Другое"])
        self.type_filter.currentTextChanged.connect(lambda: self.proxy_model.set_filter(2,
                                                                                        self.type_filter.currentText() if self.type_filter.currentText() != "Все" else ""))
        filter_layout.addRow("Тип:", self.type_filter)
        self.shelf_filter = QComboBox()
        self.shelf_filter.addItems(["Все"] + self.manager.shelves)
        self.shelf_filter.currentTextChanged.connect(lambda: self.proxy_model.set_filter(4,
                                                                                         self.shelf_filter.currentText() if self.shelf_filter.currentText() != "Все" else ""))
        filter_layout.addRow("Стеллаж:", self.shelf_filter)
        self.rack_filter = QLineEdit()
        self.rack_filter.setPlaceholderText("Поиск по полке...")
        self.rack_filter.textChanged.connect(lambda: self.proxy_model.set_filter(5, self.rack_filter.text()))
        filter_layout.addRow("Полка:", self.rack_filter)
        self.doc_number_filter = QLineEdit()
        self.doc_number_filter.setPlaceholderText("Поиск по номеру документа...")
        self.doc_number_filter.textChanged.connect(
            lambda: self.proxy_model.set_filter(6, self.doc_number_filter.text()))
        filter_layout.addRow("Номер документа:", self.doc_number_filter)
        self.category_filter = QComboBox()
        self.category_filter.addItems([
            "Все",
            "ТС - Теплосеть (отопление + ГВС или перегретая вода)",
            "ВО - Хоз. бытовая канализация",
            "ВС - Водоснабжение (ХВС)",
            "ЛК - Ливневая канализация",
            "УУТЭ - Узел учета тепловой энергии",
            "УУХВС - Узел учета холодного водоснабжения",
            "Не указана"
        ])
        self.category_filter.currentTextChanged.connect(lambda: self.proxy_model.set_filter(9,
                                                                                            self.category_filter.currentText().split(
                                                                                                " -")[
                                                                                                0].strip() if self.category_filter.currentText() not in [
                                                                                                "Все",
                                                                                                "Не указана"] else ""))
        filter_layout.addRow("Категория:", self.category_filter)
        clear_filters_btn = AnimatedButton("🗑 Сбросить фильтры")
        clear_filters_btn.clicked.connect(self._clear_filters)
        filter_layout.addRow(clear_filters_btn)
        filter_group.setLayout(filter_layout)
        main_layout.addWidget(filter_group)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        self.table = QTableView()
        self.table.setModel(self.proxy_model)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.doubleClicked.connect(self._on_table_double_click)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.tree = QTreeView()
        self.tree.setModel(self.tree_model)
        self.tree.setAlternatingRowColors(True)
        self.tree.doubleClicked.connect(self._on_tree_double_click)
        splitter.addWidget(self.table)
        splitter.addWidget(self.tree)
        splitter.setSizes([1000, 500])
        main_layout.addWidget(splitter)

    def _clear_filters(self):
        """Сброс всех фильтров."""
        self.name_filter.clear()
        self.type_filter.setCurrentText("Все")
        self.shelf_filter.setCurrentText("Все")
        self.rack_filter.clear()
        self.doc_number_filter.clear()
        self.category_filter.setCurrentText("Все")
        self.proxy_model.clear_filters()

    def refresh_data(self):
        """Обновление данных в таблице и дереве."""
        try:
            self.model.refresh_cache()
            self._populate_tree()
            self.table.resizeColumnsToContents()
            logger.info("Данные в окне просмотра обновлены")
        except Exception as e:
            logger.error(f"Ошибка при обновлении данных: {e}")
            QMessageBox.critical(self, "Ошибка", str(e))

    def _populate_tree(self):
        """Заполнение дерева иерархии."""
        self.tree_model.removeRows(0, self.tree_model.rowCount())
        elements = self.model.all_elements
        root_items = {}
        for el_id, el in elements.items():
            item = QStandardItem(f"{el['Тип']}: {el['Название']}")
            item.setData(el_id, Qt.ItemDataRole.UserRole)
            parent_id = el.get("Родитель ID")
            if not parent_id:
                root_items[el_id] = item
                self.tree_model.appendRow(item)
            else:
                parent_item = self._find_item_by_id(parent_id, self.tree_model.invisibleRootItem())
                if parent_item:
                    parent_item.appendRow(item)
        self.tree.expandAll()

    def _find_item_by_id(self, el_id, parent_item):
        """Поиск элемента в дереве по ID."""
        for row in range(parent_item.rowCount()):
            item = parent_item.child(row)
            if item.data(Qt.ItemDataRole.UserRole) == el_id:
                return item
            found = self._find_item_by_id(el_id, item)
            if found:
                return found
        return None

    def _get_type_icon(self, el_type):
        """Получение иконки для типа элемента."""
        icons = {
            "Документ": "📄",
            "Коробка": "📦",
            "Папка": "📁",
            "Другое": "🗂"
        }
        return icons.get(el_type, "🗂")

    @lru_cache(maxsize=1000)
    def _get_cached_parent_name(self, parent_id):
        """Кэшированное получение имени родителя."""
        if not parent_id:
            return "Корень (нет родителя)"
        parent = self.model.all_elements.get(parent_id)
        return f"{parent['Тип']}: {parent['Название']}" if parent else "Не найден"

    def export_to_excel(self):
        """Экспорт данных в Excel."""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.critical(self, "Ошибка", "Модуль openpyxl не установлен")
            return
        try:
            app_dir = get_app_dir()
            exports_dir = os.path.join(app_dir, 'exports')
            os.makedirs(exports_dir, exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.title = "Архив"
            headers = self.model.headers
            ws.append(headers)
            for row in self.model.elements:
                ws.append(row)
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(fill_type="solid", fgColor="2196F3")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for col in ws.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        try:
                            cell_len = max(len(str(s)) for s in str(cell.value).split("\n"))
                            max_length = max(max_length, cell_len)
                        except:
                            pass
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                ws.column_dimensions[column_letter].width = min(max_length + 3, 60)
            tab = Table(displayName="ArchiveTable", ref=ws.dimensions)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
            file_name = os.path.join(exports_dir, "archive_export.xlsx")
            wb.save(file_name)
            QMessageBox.information(self, "✅ Успех", f"Данные экспортированы в:\n{file_name}")
            logger.info(f"Экспорт Excel: {file_name}")
        except Exception as e:
            logger.error(f"Ошибка экспорта Excel: {e}")
            QMessageBox.critical(self, "❌ Ошибка", f"Не удалось экспортировать:\n{str(e)}")

    def export_to_csv(self):
        """Экспорт данных в CSV без ID и с префиксами категорий."""
        try:
            app_dir = get_app_dir()
            exports_dir = os.path.join(app_dir, 'exports')
            os.makedirs(exports_dir, exist_ok=True)

            file_name = os.path.join(exports_dir, "archive_export.csv")

            with open(file_name, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)

                # Заголовки без ID
                csv_headers = ["Название", "Тип", "Родитель ID", "Стеллаж", "Полка",
                              "Номер документа", "Дата подписания", "Расположение", "Категория"]
                writer.writerow(csv_headers)

                # Данные с префиксами категорий и без "коробка" в конце
                for row in self.model.elements:
                    # Пропускаем ID (первый столбец)
                    data_row = list(row[1:])  # Начинаем со второго столбца

                    # Получаем категорию и добавляем префикс
                    category = data_row[-1] if data_row else ""  # Категория в последнем столбце
                    prefix = ""
                    if category:
                        # Преобразуем коды категорий в префиксы
                        category_codes = category.split(",")
                        prefixes = []
                        for code in category_codes:
                            code = code.strip()
                            if code == "ТС":
                                prefixes.append("ТО")
                            elif code == "ВО":
                                prefixes.append("ВО")
                            elif code == "ВС":
                                prefixes.append("ВС")
                            elif code == "ЛК":
                                prefixes.append("ЛК")
                            elif code == "УУТЭ":
                                prefixes.append("УУТЭ")
                            elif code == "УУХВС":
                                prefixes.append("УУХВС")
                        if prefixes:
                            prefix = "/".join(prefixes) + ": "

                    # Убираем "коробка" из расположения (столбец 6 - Расположение)
                    if len(data_row) > 6:
                        location = data_row[6]
                        if location and "Коробка" in location:
                            # Убираем упоминание коробки из пути
                            parts = location.split(" / ")
                            filtered_parts = [part for part in parts if not part.startswith("Коробка")]
                            data_row[6] = " / ".join(filtered_parts)

                    # Добавляем префикс к названию (столбец 0)
                    if data_row and prefix:
                        data_row[0] = prefix + str(data_row[0])

                    writer.writerow(data_row)

            QMessageBox.information(self, "✅ Успех", f"Данные экспортированы в:\n{file_name}")
            logger.info(f"Экспорт CSV: {file_name}")
        except Exception as e:
            logger.error(f"Ошибка экспорта CSV: {e}")
            QMessageBox.critical(self, "❌ Ошибка", f"Не удалось экспортировать:\n{str(e)}")

    def _start_web_server_if_needed(self):
        """Запуск веб-сервера в отдельном потоке, если он еще не запущен."""
        global _web_server_thread, _web_server_running

        if _web_server_running:
            return

        try:
            from web_server import start_web_server

            def run_server():
                global _web_server_running
                try:
                    _web_server_running = True
                    start_web_server(port=8080)
                except Exception as e:
                    logger.error(f"Ошибка веб-сервера: {e}")
                    _web_server_running = False

            _web_server_thread = threading.Thread(target=run_server, daemon=True)
            _web_server_thread.start()
            logger.info("Веб-сервер запущен для QR-кодов")
        except Exception as e:
            logger.error(f"Не удалось запустить веб-сервер: {e}")

    def print_labels(self):
        """Печать наклеек на коробки."""
        if not REPORTLAB_AVAILABLE:
            QMessageBox.critical(self, "Ошибка", "Модули reportlab и qrcode не установлены")
            return

        try:
            # Проверка BASE_URL
            base_url = os.environ.get('BASE_URL') or os.environ.get('VERCEL_URL') or 'http://localhost:8080'

            # Если localhost, предлагаем ввести URL
            if 'localhost' in base_url or '127.0.0.1' in base_url:
                url, ok = QInputDialog.getText(
                    self,
                    "URL для QR-кодов",
                    "Введите URL вашего развернутого сайта:\n"
                    "(например: https://your-app.railway.app)\n\n"
                    "Оставьте пустым для использования localhost:",
                    text=base_url
                )

                if ok and url and url.strip():
                    # Сохраняем в переменную окружения для текущей сессии
                    base_url = url.strip()
                    if not base_url.startswith('http'):
                        base_url = f'https://{base_url}'
                    os.environ['BASE_URL'] = base_url
                elif not ok:
                    # Пользователь отменил
                    return
                else:
                    # Пользователь оставил пустым или localhost
                    reply = QMessageBox.warning(
                        self, "⚠️ Предупреждение",
                        "QR-коды будут содержать localhost, который не работает на телефоне!\n\n"
                        "Продолжить с localhost?",
                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                        QMessageBox.StandardButton.No
                    )
                    if reply == QMessageBox.StandardButton.No:
                        return

            # Сохраняем BASE_URL для использования в QR-кодах
            self.base_url = base_url
            if not self.base_url.startswith('http'):
                self.base_url = f'https://{self.base_url}'

            # Запустить веб-сервер для QR-кодов
            self._start_web_server_if_needed()

            # Получить все коробки из текущего фильтра
            boxes_data = []
            for row_data in self.model.elements:
                if len(row_data) >= 3 and row_data[2] == "Коробка":  # Тип == Коробка
                    box_info = {
                        "ID": row_data[0],
                        "Название": row_data[1],
                        "Стеллаж": row_data[4] or "",
                        "Полка": row_data[5] or "",
                        "Категория": row_data[9] or ""
                    }
                    boxes_data.append(box_info)

            if not boxes_data:
                QMessageBox.information(self, "Информация", "Нет коробок для печати в текущем фильтре")
                return

            # Открыть диалог настроек печати
            dialog = PrintLabelsDialog(boxes_data, self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                settings = dialog.get_print_settings()
                self._generate_labels_pdf(settings)

        except Exception as e:
            logger.error(f"Ошибка печати наклеек: {e}")
            QMessageBox.critical(self, "❌ Ошибка", f"Не удалось подготовить наклейки:\n{str(e)}")

    def _generate_labels_pdf(self, settings):
        """Генерация PDF с наклейками."""
        try:
            app_dir = get_app_dir()
            exports_dir = os.path.join(app_dir, 'exports')
            os.makedirs(exports_dir, exist_ok=True)

            filename = os.path.join(exports_dir, "box_labels.pdf")

            # Размеры страницы A4
            page_width, page_height = A4
            cols = settings["layout"]["cols"]
            rows = settings["layout"]["rows"]

            # Расчет размеров наклейки с ПРАВИЛЬНЫМИ отступами
            # Уменьшаем margin для максимального использования пространства
            margin_x = 0.5 * cm  # Горизонтальный отступ от края
            margin_y = 0.5 * cm  # Вертикальный отступ от края

            # Расстояние между наклейками
            gap_x = 0.2 * cm  # Промежуток между колонками
            gap_y = 0.2 * cm  # Промежуток между рядами

            # Доступное пространство для наклеек
            available_width = page_width - 2 * margin_x
            available_height = page_height - 2 * margin_y

            # Размер одной наклейки
            label_width = (available_width - (cols - 1) * gap_x) / cols
            label_height = (available_height - (rows - 1) * gap_y) / rows

            # Создание PDF
            c = canvas.Canvas(filename, pagesize=A4)
            c.setTitle("Наклейки на коробки")

            # Регистрация шрифта с поддержкой кириллицы
            try:
                from reportlab.pdfbase.cidfonts import UnicodeCIDFont
                pdfmetrics.registerFont(UnicodeCIDFont('HeiseiMin-W3'))
            except:
                try:
                    pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))
                except:
                    logger.warning("Не удалось зарегистрировать шрифт с поддержкой кириллицы, используется Helvetica")

            boxes = settings["selected_boxes"]
            format_type = settings["format_type"]
            custom_options = settings.get("custom_options", {})

            label_count = 0
            page_count = 0
            qr_count = 0

            logger.info(f"Начало генерации PDF с {len(boxes)} наклейками")
            logger.info(f"Размер наклейки: {label_width/cm:.2f}x{label_height/cm:.2f} см")
            logger.info(f"Сетка: {cols}x{rows}")

            for box in boxes:
                if label_count >= cols * rows:
                    # Новая страница
                    c.showPage()
                    page_count += 1
                    label_count = 0

                # Расчет позиции наклейки в сетке
                col = label_count % cols
                row = label_count // cols

                # ИСПРАВЛЕННАЯ формула позиции с учётом промежутков
                x = margin_x + col * (label_width + gap_x)
                y = page_height - margin_y - (row + 1) * label_height - row * gap_y

                # Генерация содержимого наклейки
                qr_added = self._draw_label_content(c, box, x, y, label_width, label_height,
                                    format_type, custom_options)
                if qr_added:
                    qr_count += 1

                label_count += 1

            c.save()

            # Получаем BASE_URL для отображения
            display_url = self.base_url if hasattr(self, 'base_url') and self.base_url else (
                os.environ.get('BASE_URL') or os.environ.get('VERCEL_URL') or 'http://localhost:8080'
            )
            if display_url and not display_url.startswith('http'):
                display_url = f'https://{display_url}'

            url_warning = ""
            if 'localhost' in display_url or '127.0.0.1' in display_url:
                url_warning = "\n\n⚠️ Внимание: QR-коды содержат localhost!\nОни не будут работать на телефоне."

            QMessageBox.information(
                self, "✅ Успех",
                f"Наклейки сгенерированы: {filename}\n"
                f"Страниц: {page_count + 1}, Наклеек: {len(boxes)}, QR-кодов: {qr_count}\n"
                f"Размер наклейки: {label_width/cm:.1f}x{label_height/cm:.1f} см\n"
                f"URL для QR-кодов: {display_url}{url_warning}"
            )
            logger.info(f"Сгенерированы наклейки: {filename}, QR-кодов: {qr_count} из {len(boxes)}")

        except Exception as e:
            logger.error(f"Ошибка генерации PDF: {e}")
            QMessageBox.critical(self, "❌ Ошибка", f"Не удалось сгенерировать PDF:\n{str(e)}")

    def _draw_label_content(self, canvas, box, x, y, width, height, format_type, custom_options):
        """Отрисовка содержимого наклейки с двумя ячейками."""
        try:
            # Определяем шрифт для кириллицы
            try:
                cyrillic_font = "HeiseiMin-W3"
                cyrillic_font_bold = "HeiseiMin-W3"
            except:
                cyrillic_font = "Helvetica"
                cyrillic_font_bold = "Helvetica-Bold"

            # === ДЕЛИМ НАКЛЕЙКУ НА ДВЕ ЯЧЕЙКИ ===
            # Левая ячейка: 60% ширины для текста
            # Правая ячейка: 40% ширины для QR-кода
            text_cell_width = width * 0.6
            qr_cell_width = width * 0.4

            # Рисуем внешнюю рамку наклейки
            canvas.setLineWidth(1.5)
            canvas.setStrokeColorRGB(0, 0, 0)
            canvas.rect(x, y, width, height)

            # Рисуем вертикальную разделительную линию между ячейками
            canvas.setLineWidth(1)
            canvas.setStrokeColorRGB(0.5, 0.5, 0.5)
            canvas.line(x + text_cell_width, y, x + text_cell_width, y + height)

            # === ЛЕВАЯ ЯЧЕЙКА: ТЕКСТОВАЯ ИНФОРМАЦИЯ ===
            text_padding = 0.3 * cm
            text_x = x + text_padding
            text_y = y + height - text_padding
            text_max_width = text_cell_width - 2 * text_padding

            # Размеры шрифтов
            font_size_name = max(10, min(14, int(height / 2.5)))
            font_size_info = max(8, min(11, int(height / 3.5)))
            line_spacing = font_size_info * 1.5

            current_y = text_y

            # 1. НАЗВАНИЕ (в цветной рамке)
            if format_type == "brief" or (format_type == "custom" and custom_options.get("show_name", True)):
                name = box["Название"]

                # Разбиваем длинное название на строки
                try:
                    canvas.setFont(cyrillic_font_bold, font_size_name)
                    words = name.split()
                    lines = []
                    current_line = []

                    for word in words:
                        test_line = " ".join(current_line + [word])
                        test_width = canvas.stringWidth(test_line, cyrillic_font_bold, font_size_name)

                        if test_width <= text_max_width:
                            current_line.append(word)
                        else:
                            if current_line:
                                lines.append(" ".join(current_line))
                                current_line = [word]
                            else:
                                lines.append(word[:int(text_max_width / (font_size_name * 0.6))] + "...")
                                current_line = []

                    if current_line:
                        lines.append(" ".join(current_line))

                    # Ограничиваем 2 строками
                    lines = lines[:2]

                except:
                    canvas.setFont("Helvetica-Bold", font_size_name)
                    lines = [name[:int(text_max_width / (font_size_name * 0.6))]]

                # Рисуем заголовок с синим фоном
                header_height = len(lines) * (font_size_name * 1.3) + 0.3 * cm
                canvas.setFillColorRGB(0.2, 0.5, 0.9)  # Синий фон
                canvas.setStrokeColorRGB(0.2, 0.5, 0.9)
                canvas.rect(text_x - 0.15*cm, current_y - header_height + 0.25*cm,
                        text_max_width + 0.3*cm, header_height, fill=1, stroke=1)

                # Белый текст на синем фоне
                canvas.setFillColorRGB(1, 1, 1)
                for line in lines:
                    canvas.drawString(text_x, current_y - font_size_name, line)
                    current_y -= font_size_name * 1.3

                current_y -= 0.4 * cm

            # Восстанавливаем черный цвет
            canvas.setFillColorRGB(0, 0, 0)

            # 2. РАСПОЛОЖЕНИЕ (с иконками)
            try:
                canvas.setFont(cyrillic_font, font_size_info)
            except:
                canvas.setFont("Helvetica", font_size_info)

            if format_type == "brief" or (format_type == "custom" and custom_options.get("show_location", True)):
                if box.get("Стеллаж"):
                    location_text = f"📚 Стеллаж: {box['Стеллаж']}"
                    if canvas.stringWidth(location_text, cyrillic_font, font_size_info) > text_max_width:
                        location_text = f"Ст: {box['Стеллаж']}"

                    # Фон для расположения
                    canvas.setFillColorRGB(0.95, 0.95, 0.95)
                    canvas.setStrokeColorRGB(0.8, 0.8, 0.8)
                    canvas.rect(text_x - 0.1*cm, current_y - font_size_info - 0.15*cm,
                            text_max_width + 0.2*cm, font_size_info + 0.3*cm, fill=1, stroke=1)

                    canvas.setFillColorRGB(0, 0, 0)
                    canvas.drawString(text_x, current_y - font_size_info, location_text)
                    current_y -= line_spacing

                if box.get("Полка"):
                    shelf_text = f"📊 Полка: {box['Полка']}"
                    if canvas.stringWidth(shelf_text, cyrillic_font, font_size_info) > text_max_width:
                        shelf_text = f"П: {box['Полка']}"

                    # Фон для полки
                    canvas.setFillColorRGB(0.95, 0.95, 0.95)
                    canvas.setStrokeColorRGB(0.8, 0.8, 0.8)
                    canvas.rect(text_x - 0.1*cm, current_y - font_size_info - 0.15*cm,
                            text_max_width + 0.2*cm, font_size_info + 0.3*cm, fill=1, stroke=1)

                    canvas.setFillColorRGB(0, 0, 0)
                    canvas.drawString(text_x, current_y - font_size_info, shelf_text)
                    current_y -= line_spacing

                current_y -= 0.2 * cm

            # 3. КАТЕГОРИЯ
            if format_type == "full" or (format_type == "custom" and custom_options.get("show_category", True)):
                category = box.get("Категория", "")
                if category:
                    category_codes = []
                    for cat in category.split(","):
                        cat = cat.strip()
                        if cat in ["ТС", "ВО", "ВС", "ЛК", "УУТЭ", "УУХВС"]:
                            category_codes.append(cat)

                    if category_codes:
                        cat_text = "🔧 " + "/".join(category_codes)
                        canvas.setFont(cyrillic_font, font_size_info - 1)
                        canvas.setFillColorRGB(0.3, 0.3, 0.3)

                        if canvas.stringWidth(cat_text, cyrillic_font, font_size_info - 1) <= text_max_width:
                            canvas.drawString(text_x, current_y - font_size_info, cat_text)

            # === ПРАВАЯ ЯЧЕЙКА: QR-КОД ===
            qr_added = False
            if (format_type == "full" or format_type == "brief" or
                (format_type == "custom" and custom_options.get("show_qr", True))):

                # QR-код занимает всю правую ячейку с небольшими отступами
                qr_padding = 0.3 * cm
                qr_available_space = min(qr_cell_width - 2 * qr_padding, height - 2 * qr_padding)
                qr_size = max(qr_available_space, 2.0 * cm)

                # Центрируем QR-код в правой ячейке
                qr_cell_x = x + text_cell_width
                qr_x = qr_cell_x + (qr_cell_width - qr_size) / 2
                qr_y = y + (height - qr_size) / 2

                # Белый фон для QR-кода
                canvas.setFillColorRGB(1, 1, 1)
                canvas.setStrokeColorRGB(0.9, 0.9, 0.9)
                canvas.rect(qr_x - 0.15*cm, qr_y - 0.15*cm,
                        qr_size + 0.3*cm, qr_size + 0.3*cm, fill=1, stroke=1)

                qr_added = self._add_qr_code(canvas, box["ID"], qr_x, qr_y, qr_size)

                # Подпись под QR
                if qr_added:
                    canvas.setFillColorRGB(0.5, 0.5, 0.5)
                    canvas.setFont("Helvetica", 7)
                    qr_label = "Сканируй"
                    label_width = canvas.stringWidth(qr_label, "Helvetica", 7)
                    canvas.drawString(qr_x + (qr_size - label_width) / 2,
                                    qr_y - 0.35*cm, qr_label)

            return qr_added

        except Exception as e:
            logger.error(f"Ошибка отрисовки наклейки: {e}")
            # В случае ошибки рисуем базовую информацию
            try:
                canvas.setFont("Helvetica", 10)
                canvas.drawString(x + 0.3*cm, y + height - 0.5*cm, box.get("Название", "Ошибка"))
            except:
                pass
            return False

    def _add_qr_code(self, canvas, box_id, x, y, size):
        """Добавление QR-кода на наклейку."""
        try:
            # Генерация QR-кода с URL
            # Используем сохраненный BASE_URL или переменную окружения
            if hasattr(self, 'base_url') and self.base_url:
                base_url = self.base_url
            else:
                base_url = os.environ.get('BASE_URL') or os.environ.get('VERCEL_URL') or 'http://localhost:8080'
                if base_url and not base_url.startswith('http'):
                    base_url = f'https://{base_url}'
            url = f"{base_url}/box/{box_id}"

            # Создание QR-кода с улучшенными настройками
            qr = qrcode.QRCode(
                version=None,  # Автоматический выбор версии
                error_correction=qrcode.constants.ERROR_CORRECT_M,  # Средняя коррекция ошибок
                box_size=8,  # Меньший размер пикселей для лучшего качества при печати
                border=2,  # Граница
            )
            qr.add_data(url)
            qr.make(fit=True)

            # Создание изображения QR-кода
            img = qr.make_image(fill_color="black", back_color="white")

            # Сохранение в BytesIO
            img_buffer = io.BytesIO()
            img.save(img_buffer, format='PNG')
            img_buffer.seek(0)

            # Добавление на canvas ReportLab
            # ReportLab может работать с BytesIO напрямую
            from reportlab.lib.utils import ImageReader
            img_reader = ImageReader(img_buffer)
            canvas.drawImage(img_reader, x, y, width=size, height=size, mask='auto')

            logger.debug(f"QR-код добавлен для коробки {box_id} по адресу {url}")
            return True

        except Exception as e:
            logger.error(f"Ошибка генерации QR-кода для коробки {box_id}: {e}")
            # В случае ошибки рисуем рамку вместо QR-кода
            try:
                canvas.setStrokeColorRGB(0.5, 0.5, 0.5)
                canvas.setLineWidth(1)
                canvas.rect(x, y, size, size)
                canvas.drawString(x + size/4, y + size/2, "QR")
            except:
                pass
            return False

    def migrate_data(self):
        """Миграция данных из JSON."""
        manager = None
        try:
            manager = DataManager(self.db_file)
            app_dir = get_app_dir()
            json_file = os.path.join(app_dir, 'elements.json')
            if manager.migrate_from_json(json_file):
                QMessageBox.information(self, "✅ Успех", f"Данные мигрированы в:\n{self.db_file}")
                self.refresh_data()
            else:
                QMessageBox.warning(self, "⚠ Предупреждение", "Файл elements.json не найден")
        except Exception as e:
            logger.error(f"Ошибка миграции: {e}")
            QMessageBox.critical(self, "❌ Ошибка", f"Не удалось выполнить миграцию:\n{str(e)}")
        finally:
            if manager:
                manager.close()

    def _on_tree_double_click(self, index):
        """Обработка двойного клика по дереву."""
        try:
            el_id = index.data(Qt.ItemDataRole.UserRole)
            if el_id:
                element = self.model.all_elements.get(el_id)
                if element:
                    self._show_element_details(element, el_id)
        except Exception as e:
            logger.error(f"Ошибка обработки клика по дереву: {e}")

    def _on_table_double_click(self, index):
        """Обработка двойного клика по таблице."""
        try:
            if not index.isValid():
                return
            source_index = self.proxy_model.mapToSource(index)
            row = source_index.row()
            el_id = self.model.elements[row][0]
            element = self.model.all_elements.get(el_id)
            if element:
                self._show_element_details(element, el_id)
        except Exception as e:
            logger.error(f"Ошибка обработки клика по таблице: {e}")

    def _show_element_details(self, element, el_id):
        """Отображение детальной информации об элементе."""
        icon = self._get_type_icon(element['Тип'])
        category = element.get('Категория', '')
        category_full = get_category_description(category) if category else 'Не указана'
        details = (
            f"ID: {el_id}\n"
            f"{icon} Тип: {element['Тип']}\n"
            f"📝 Название: {element['Название']}\n"
            f"📂 Родитель: {self._get_cached_parent_name(element.get('Родитель ID'))}\n"
            f"📚 Стеллаж: {element.get('Стеллаж') or 'Не указан'}\n"
            f"📊 Полка: {element.get('Полка') or 'Не указана'}\n"
            f"🔢 Номер документа: {element.get('Номер документа') or 'Не указан'}\n"
            f"📅 Дата подписания: {element.get('Дата подписания') or 'Не указана'}\n"
            f"🔧 Категория: {category_full}"
        )
        QMessageBox.information(self, f"{icon} Детали элемента", details)

    def _get_category_full_name(self, category_code):
        """Получение полного названия категории по коду."""
        return get_category_description(category_code)

    def back_to_menu(self):
        """Возврат в главное меню."""
        if self.main_menu:
            main_menu = self.main_menu()
            if main_menu:
                main_menu.show()
        self.close()

    def _animate_window(self):
        """Анимация появления окна."""
        self.setWindowOpacity(0)
        self.animation = QPropertyAnimation(self, b"windowOpacity")
        self.animation.setDuration(400)
        self.animation.setStartValue(0)
        self.animation.setEndValue(1)
        self.animation.setEasingCurve(QEasingCurve.Type.OutCubic)
        self.animation.start()

    def closeEvent(self, event):
        """Обработка закрытия окна."""
        try:
            self._get_cached_parent_name.cache_clear()
            logger.info("ViewWindow закрыто")
        except Exception as e:
            logger.error(f"Ошибка при закрытии ViewWindow: {e}")
        finally:
            super().closeEvent(event)
